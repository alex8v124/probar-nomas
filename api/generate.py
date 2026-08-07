from flask import Flask, request, send_file, jsonify
import io
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import base64
from io import BytesIO
from PIL import Image as PILImage
import urllib.request
import json

app = Flask(__name__)

@app.route('/api/check', methods=['GET'])
def check_limit():
    try:
        url = 'https://api.counterapi.dev/v1/alex_probar_nomas/generador_excel'
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req) as response:
            data = json.loads(response.read().decode())
            count = data.get('count', 0)
            if count >= 100:
                return jsonify({"blocked": True, "count": count})
            return jsonify({"blocked": False, "count": count})
    except Exception as e:
        return jsonify({"blocked": False, "count": 0})

@app.route('/api/generate', methods=['POST'])
def generate():
    try:
        # Incrementar contador y verificar bloqueo
        try:
            url = 'https://api.counterapi.dev/v1/alex_probar_nomas/generador_excel/up'
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req) as response:
                count_data = json.loads(response.read().decode())
                if count_data.get('count', 0) > 100:
                    return jsonify({"error": "Límite máximo alcanzado"}), 403
        except Exception:
            pass # Si la API del contador falla, permitimos continuar
            
        data = request.json
        client_name = data.get('client_name', '')
        phone = data.get('phone', '')
        address = data.get('address', '')
        city = data.get('city', '')
        zip_code = data.get('zip_code', '')
        items = data.get('items', [])
        grand_total = data.get('grand_total', 0)

        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Estilos generales
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        white_font_bold = Font(color='FFFFFF', bold=True, size=12)
        large_white_font_bold = Font(color='FFFFFF', bold=True, size=24)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Configurar anchos de columna
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 55
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 22
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15

        # Aplicar fondo y borde a todas las celdas de fila 1 y 2 antes de hacer merge
        for r in range(1, 3):
            for c in range(1, 10):
                cell = ws.cell(row=r, column=c)
                cell.fill = red_fill
                cell.border = thin_border

        # Fila 1: Cliente y Telefono
        ws.row_dimensions[1].height = 60
        ws.merge_cells('A1:F1')
        ws['A1'] = f"NOMBRE DEL CLIENTE (CUSTOMER NAME) : {client_name}"
        ws['A1'].fill = red_fill
        ws['A1'].font = large_white_font_bold
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.merge_cells('G1:I1')
        ws['G1'] = f"TELEFONO: {phone}"
        ws['G1'].fill = red_fill
        ws['G1'].font = large_white_font_bold
        ws['G1'].alignment = Alignment(horizontal='center', vertical='center')

        # Fila 2: Dirección, Ciudad, Codigo Postal
        ws.row_dimensions[2].height = 60
        ws.merge_cells('A2:C2')
        ws['A2'] = f"DIRECCIÓN: {address}"
        ws['A2'].fill = red_fill
        ws['A2'].font = large_white_font_bold
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('D2:F2')
        ws['D2'] = f"CIUDAD: {city}"
        ws['D2'].fill = red_fill
        ws['D2'].font = large_white_font_bold
        ws['D2'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('G2:I2')
        ws['G2'] = f"CODIGO POSTAL: {zip_code}"
        ws['G2'].fill = red_fill
        ws['G2'].font = large_white_font_bold
        ws['G2'].alignment = Alignment(horizontal='center', vertical='center')

        # Fila 3: Cabeceras sin paréntesis
        headers = ["Nº", "PRODUCTO", "PEDIDO", "DESCRIPCIÓN DE PRODUCTO", "TOTAL", "ENLACE", "PEDIDO TOTAL", "PRECIO UNIDAD", "PRECIO TOTAL"]
        ws.row_dimensions[3].height = 40
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.fill = red_fill
            cell.font = white_font_bold
            cell.alignment = center_align
            cell.border = thin_border

        # Productos
        current_row = 4
        for idx, item in enumerate(items, 1):
            ws.row_dimensions[current_row].height = 350
            
            # Textos
            ws.cell(row=current_row, column=1, value=idx).alignment = center_align
            ws.cell(row=current_row, column=4, value=item.get('desc', '')).alignment = center_align
            ws.cell(row=current_row, column=5, value=item.get('total', '')).alignment = center_align
            
            link_val = item.get('link', '')
            link_cell = ws.cell(row=current_row, column=6, value=link_val)
            link_cell.alignment = center_align
            if link_val:
                # Nos aseguramos de que tenga un esquema de url válido para excel
                if not (link_val.startswith('http://') or link_val.startswith('https://')):
                    link_cell.hyperlink = 'https://' + link_val
                else:
                    link_cell.hyperlink = link_val
                link_cell.font = Font(color="0000FF", underline="single")
                link_cell.value = link_val  # Set value explicitly for display
            # Formato de Yuanes para contabilidad
            yuan_format = '_-"¥"* #,##0.00_-;\\-"¥"* #,##0.00_-;_-"¥"* "-"??_-;_-@_-'
            
            ws.cell(row=current_row, column=7, value=item.get('total_pedido', 0)).alignment = center_align
            # Eliminado formato de yuanes para columna 7 (cantidad)
            
            ws.cell(row=current_row, column=8, value=item.get('precio_unidad', 0)).alignment = center_align
            ws.cell(row=current_row, column=8).number_format = yuan_format
            
            ws.cell(row=current_row, column=9, value=item.get('precio_total', 0)).alignment = center_align
            ws.cell(row=current_row, column=9).number_format = yuan_format

            # Bordes para todas las celdas
            for col in range(1, 10):
                ws.cell(row=current_row, column=col).border = thin_border

            # Función helper para insertar imágenes
            def insert_image(base64_str, col_idx, col_width_units):
                if not base64_str: return
                try:
                    if "," in base64_str:
                        base64_str = base64_str.split(",")[1]
                    img_data = base64.b64decode(base64_str)
                    
                    pil_img = PILImage.open(BytesIO(img_data))
                    
                    # Col width in pixels (aprox)
                    col_width_px = col_width_units * 7.5
                    row_height_px = 350 * 1.33
                    
                    # Max dimensions (94% of cell)
                    max_w = col_width_px * 0.94
                    max_h = row_height_px * 0.94
                    
                    # Calcular tamaño proporcional
                    img_w, img_h = pil_img.size
                    ratio = min(max_w/img_w, max_h/img_h)
                    
                    final_w = int(img_w * ratio)
                    final_h = int(img_h * ratio)
                    
                    # Convert to RGB if needed to avoid openpyxl errors with PNGs with alpha
                    if pil_img.mode in ('RGBA', 'P'):
                        pil_img = pil_img.convert('RGB')
                        
                    # Save into a clean BytesIO
                    clean_io = BytesIO()
                    pil_img.save(clean_io, format='JPEG')
                    xl_img = OpenpyxlImage(clean_io)
                    
                    xl_img.width = final_w
                    xl_img.height = final_h
                    
                    # Centrado usando celdas y offsets
                    col_offset_px = int((col_width_px - final_w) / 2)
                    row_offset_px = int((row_height_px - final_h) / 2)
                    
                    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                    from openpyxl.drawing.xdr import XDRPositiveSize2D
                    
                    marker = AnchorMarker(col=col_idx - 1, colOff=col_offset_px * 9525, 
                                          row=current_row - 1, rowOff=row_offset_px * 9525)
                    size = XDRPositiveSize2D(final_w * 9525, final_h * 9525)
                    xl_img.anchor = OneCellAnchor(_from=marker, ext=size)
                    
                    ws.add_image(xl_img)
                except Exception as e:
                    print(f"Error procesando imagen: {e}")

            insert_image(item.get('prod_img'), 2, 45) # Col B (2)
            insert_image(item.get('mix_img'), 3, 55)  # Col C (3)
            
            current_row += 1

        # Fila de Total
        ws.merge_cells(f'D{current_row}:H{current_row}')
        cell_total_text = ws.cell(row=current_row, column=4, value="PRECIO TOTAL")
        cell_total_text.fill = red_fill
        cell_total_text.font = white_font_bold
        cell_total_text.alignment = Alignment(horizontal='right', vertical='center')
        
        yuan_format = '_-"¥"* #,##0.00_-;\\-"¥"* #,##0.00_-;_-"¥"* "-"??_-;_-@_-'
        cell_total_val = ws.cell(row=current_row, column=9, value=grand_total)
        cell_total_val.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        cell_total_val.font = Font(bold=True)
        cell_total_val.alignment = center_align
        cell_total_val.number_format = yuan_format

        for col in range(4, 10):
            ws.cell(row=current_row, column=col).border = thin_border

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        return send_file(
            out,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='PEDIDO_1688.xlsx'
        )
    except Exception as e:
        return {"error": str(e)}, 500

if __name__ == '__main__':
    app.run(debug=True)
